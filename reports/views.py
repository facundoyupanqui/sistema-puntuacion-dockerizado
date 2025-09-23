import io
from datetime import timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Sum, F, Count
from django.http import FileResponse, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from datetime import datetime

from activities.models.activity import Activity
from users.models.user import User
from teams.models import Team
from .models.period import Period
from .models.ranking import Ranking


def is_admin(user: User) -> bool:
    return user.is_authenticated and user.is_admin


def _get_period_range(period: str):
    today = timezone.now().date()
    if period == 'daily':
        return today, today
    if period == 'weekly':
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
        return start, end
    # biweekly: 1-15, 16-end
    if today.day <= 15:
        start = today.replace(day=1)
        end = today.replace(day=15)
    else:
        start = today.replace(day=16)
        # end of month
        next_month = (today.replace(day=28) + timedelta(days=4)).replace(day=1)
        end = next_month - timedelta(days=1)
    return start, end


@login_required
def history(request):
    period = request.GET.get('period')  # daily|weekly|biweekly or custom
    user_id = request.GET.get('user')
    team_id = request.GET.get('team')
    start = request.GET.get('start')
    end = request.GET.get('end')

    qs = Activity.objects.select_related('activity_type', 'user', 'user__team')

    if period in ('daily', 'weekly', 'biweekly'):
        start_date, end_date = _get_period_range(period)
        qs = qs.filter(date__range=(start_date, end_date))
    elif start and end:
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d').date()
            end_date = datetime.strptime(end, '%Y-%m-%d').date()
            qs = qs.filter(date__range=(start_date, end_date))
        except Exception:
            pass

    if user_id and str(user_id).isdigit():
        qs = qs.filter(user_id=user_id)
    if team_id and str(team_id).isdigit():
        qs = qs.filter(user__team_id=team_id)

    qs = qs.order_by('-date', '-created_at')

    # Estadísticas
    total_activities = qs.count()
    totals = qs.aggregate(total_points=Sum(F('activity_type__points')))
    total_points = totals['total_points'] or 0
    active_users = qs.values('user_id').distinct().count()
    distinct_days = qs.values('date').distinct().count() or 1
    daily_average = round(total_activities / distinct_days)

    context = {
        'activities': qs,
        'users': User.objects.filter(is_active=True).order_by('name'),
        'teams': Team.objects.all().order_by('name'),
        'stats': {
            'total_activities': total_activities,
            'total_points': total_points,
            'active_users': active_users,
            'daily_average': daily_average,
        },
        'selected': {
            'period': period or '',
            'user': user_id or '',
            'team': team_id or '',
            'start': start or '',
            'end': end or '',
        },
    }

    return render(request, 'reports/history.html', context)


@login_required
def export_history_excel(request):
    # Generar XLSX con openpyxl (mejor compatibilidad)
    period = request.GET.get('period')
    user_id = request.GET.get('user')
    team_id = request.GET.get('team')
    start = request.GET.get('start')
    end = request.GET.get('end')

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Historial'

    qs = Activity.objects.select_related('activity_type', 'user', 'user__team')
    if period in ('daily', 'weekly', 'biweekly'):
        start_date, end_date = _get_period_range(period)
        qs = qs.filter(date__range=(start_date, end_date))
    elif start and end:
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d').date()
            end_date = datetime.strptime(end, '%Y-%m-%d').date()
            qs = qs.filter(date__range=(start_date, end_date))
        except Exception:
            pass
    if user_id and str(user_id).isdigit():
        qs = qs.filter(user_id=user_id)
    if team_id and str(team_id).isdigit():
        qs = qs.filter(user__team_id=team_id)

    headers = ['Fecha', 'Usuario', 'Equipo', 'Actividad', 'Puntos', 'Evidencia']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2D2F3A', end_color='2D2F3A', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for a in qs:
        ws.append([
            a.date.isoformat(),
            a.user.name,
            a.user.team.name if a.user.team else '-',
            a.activity_type.name,
            a.activity_type.points,
            a.evidence or '-',
        ])

    for col in ws.columns:
        max_len = 12
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = FileResponse(bio, as_attachment=True, filename='historial_actividades.xlsx', content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return resp


@login_required
def export_history_pdf(request):
    # Exportar PDF con tabla usando fpdf2 (sin dependencias de compilación)
    period = request.GET.get('period')
    user_id = request.GET.get('user')
    team_id = request.GET.get('team')
    start = request.GET.get('start')
    end = request.GET.get('end')

    qs = Activity.objects.select_related('activity_type', 'user', 'user__team')
    if period in ('daily', 'weekly', 'biweekly'):
        start_date, end_date = _get_period_range(period)
        qs = qs.filter(date__range=(start_date, end_date))
    elif start and end:
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d').date()
            end_date = datetime.strptime(end, '%Y-%m-%d').date()
            qs = qs.filter(date__range=(start_date, end_date))
        except Exception:
            pass
    if user_id and str(user_id).isdigit():
        qs = qs.filter(user_id=user_id)
    if team_id and str(team_id).isdigit():
        qs = qs.filter(user__team_id=team_id)

    from fpdf import FPDF

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'Historial de Actividades', ln=1)

    pdf.set_font('Helvetica', 'B', 10)
    headers = ['Fecha', 'Usuario', 'Equipo', 'Actividad', 'Puntos']
    widths = [25, 45, 45, 60, 20]
    for h, w in zip(headers, widths):
        pdf.cell(w, 8, h, border=1, align='C')
    pdf.ln(8)

    pdf.set_font('Helvetica', '', 10)
    for a in qs:
        row = [
            a.date.isoformat(),
            a.user.name,
            a.user.team.name if a.user.team else '-',
            a.activity_type.name,
            str(a.activity_type.points),
        ]
        for c, w in zip(row, widths):
            pdf.cell(w, 8, c, border=1, align='C')
        pdf.ln(8)

    output = bytes(pdf.output(dest='S'))
    buffer = io.BytesIO(output)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='historial_actividades.pdf', content_type='application/pdf')


@login_required
def ranking_api(request):
    period = request.GET.get('period', 'biweekly')
    start_date, end_date = _get_period_range(period)

    leaderboard = (
        User.objects.filter(is_active=True)
        .annotate(points=Sum(
            F('activity__activity_type__points'),
            filter=(
                F('activity__date__gte') == start_date
            )
        ))
    )
    # Simpler: aggregate from Activity
    leaderboard = (
        Activity.objects.filter(date__range=(start_date, end_date))
        .values('user_id', 'user__name', 'user__team__name')
        .annotate(points=Sum(F('activity_type__points')), activities=Count('id'))
        .order_by('-points')
    )

    user_position = None
    my_points = 0
    for idx, row in enumerate(leaderboard, start=1):
        if row['user_id'] == request.user.id:
            user_position = idx
            my_points = row['points'] or 0
            break

    return HttpResponse(
        content_type='application/json',
        content=__import__('json').dumps({
            'period': period,
            'start': start_date.isoformat(),
            'end': end_date.isoformat(),
            'leaderboard': list(leaderboard),
            'kpis': {
                'my_points': my_points,
                'my_position': user_position,
                'my_activities': Activity.objects.filter(user=request.user, date__range=(start_date, end_date)).count(),
                'total_points': sum((row['points'] or 0) for row in leaderboard),
            },
        })
    )


@login_required
@user_passes_test(is_admin)
def close_biweekly(request):
    # cierra el periodo vigente (quincenal), guarda ranking y ganador
    start_date, end_date = _get_period_range('biweekly')

    period, _ = Period.objects.get_or_create(
        type=Period.BIWEEKLY,
        startDate=start_date,
        endDate=end_date,
        defaults={'is_closed': False},
    )

    if period.is_closed:
        messages.info(request, 'Este periodo ya está cerrado.')
        return render(request, 'reports/close.html', {'period': period})

    leaderboard = (
        Activity.objects.filter(date__range=(start_date, end_date))
        .values('user_id')
        .annotate(points=Sum(F('activity_type__points')), activities=Count('id'))
        .order_by('-points')
    )

    Ranking.objects.filter(period=period).delete()
    for position, row in enumerate(leaderboard, start=1):
        Ranking.objects.create(
            period=period,
            position=position,
            user_id=row['user_id'],
            total_points=row['points'] or 0,
            total_activities=row['activities'],
        )

    period.is_closed = True
    period.save()
    messages.success(request, 'Periodo quincenal cerrado y ranking publicado.')
    return render(request, 'reports/close.html', {'period': period})

# Create your views here.